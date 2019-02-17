# Merges excel files, with some caveats:
#
#	- All files must have the same sheets
#	- Cells with data take priority over empty cells
#	- Otherwise, priority goes to earlier files in the args list
#
# Tested on openpyxl 2.6.0 (latest at time of writing)

import copy, sys, time
start_time = time.time()
print("Loading openpyxl...")
import openpyxl

# ---------------------------------------------------------------------

def main():

	if len(sys.argv) < 2: # Allow a single file
		print("Usage: excel_merger.py <file1> <file2> <etc>")
		return

	workbooks = []

	for filename in sys.argv[1:]:
		print("Opening " + filename)
		workbooks.append(openpyxl.load_workbook(filename))

	if same_sheet_names(workbooks) == False:
		print("Workbooks had different sheets")
		return

	print("Merging...")
	merge(workbooks, "merged_output.xlsx")
	print("Done.")
	return

# ---------------------------------------------------------------------

def same_sheet_names(workbooks):

	expected_sheets = workbooks[0].get_sheet_names()

	for workbook in workbooks[1:]:
		if workbook.get_sheet_names() != expected_sheets:
			return False

	return True

# ---------------------------------------------------------------------

def merge(workbooks, outfilename):

	# Merge data into workbook 0 and then save it as a different filename

	sheet_names = workbooks[0].get_sheet_names()

	for name in sheet_names:
		target = workbooks[0].get_sheet_by_name(name)
		for workbook in workbooks[1:]:
			source = workbook.get_sheet_by_name(name)
			for x in range(1, source.max_column + 1):
				for y in range(1, source.max_row + 1):
					sc = source.cell(column = x, row = y)
					if sc.value is not None:
						tc = target.cell(column = x, row = y)
						if tc.value is None or (type(tc.value) is str and tc.value.strip() == ""):

							tc.value = sc.value

							if type(tc.value) is str:	# For speed reasons, only do formatting on strings:

								tc.alignment = 		copy.copy(sc.alignment)
								tc.fill = 			copy.copy(sc.fill)
								tc.font = 			copy.copy(sc.font)
								tc.border = 		copy.copy(sc.border)
								tc.number_format = 	copy.copy(sc.number_format)
								tc.protection = 	copy.copy(sc.protection)

	workbooks[0].save(outfilename)

# ---------------------------------------------------------------------

main()
print("Time elapsed: {0:.2f} seconds".format(time.time() - start_time))
