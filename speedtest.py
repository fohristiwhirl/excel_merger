import random, time
import openpyxl


def main():

	wb = openpyxl.Workbook()
	ws = wb.active
	populate(ws, 20, 5000)
	test(ws, cache = True)
	test(ws, cache = False)

	wb = openpyxl.Workbook()
	ws = wb.active
	populate(ws, 5000, 20)
	test(ws, cache = True)
	test(ws, cache = False)


def test(ws, cache = True):

	max_column = ws.max_column
	max_row = ws.max_row

	print("Data size is {} x {}, caching is {}".format(max_column, max_row, "ON" if cache else "OFF"))

	# Run the test one way...

	start_time = time.time()
	max_c_calcs = 0
	max_r_calcs = 0

	if cache:
		for x in range(1, max_column + 1):
			for y in range(1, max_row + 1):
				cell = ws.cell(column = x, row = y)
	else:

		max_c_calcs += 1
		for x in range(1, ws.max_column + 1):

			max_r_calcs += 1
			for y in range(1, ws.max_row + 1):

				cell = ws.cell(column = x, row = y)

	print("x-loop outer, time elapsed: {0:.2f} seconds, max_column ({1}): {2} calc, max_row ({3}): {4} calc".format(
		time.time() - start_time, max_column, max_c_calcs, max_row, max_r_calcs))

	# Run the test the other way...

	start_time = time.time()
	max_c_calcs = 0
	max_r_calcs = 0

	if cache:
		for y in range(1, max_row + 1):
			for x in range(1, max_column + 1):
				cell = ws.cell(column = x, row = y)
	else:

		max_r_calcs += 1
		for y in range(1, ws.max_row + 1):

			max_c_calcs += 1
			for x in range(1, ws.max_column + 1):

				cell = ws.cell(column = x, row = y)

	print("y-loop outer, time elapsed: {0:.2f} seconds, max_column ({1}): {2} calc, max_row ({3}): {4} calc".format(
		time.time() - start_time, max_column, max_c_calcs, max_row, max_r_calcs))


def populate(ws, width, height):
	for x in range(1, width + 1):
		for y in range(1, height + 1):
			if random.choice([True, False]):
				ws.cell(column = x, row = y).value = random.randint(0, 100)


main()
